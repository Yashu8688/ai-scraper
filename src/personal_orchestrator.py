import json
import logging
from typing import List, Dict, Any
from config import settings
from src.scrapers import GreenhouseScraper, AshbyScraper
from src.filters.personal_parser import filter_personal_job
from src.reporting.excel import generate_styled_excel
from src.reporting.email_client import send_email_with_report

logger = logging.getLogger(__name__)

PERSONAL_COMPANIES_PATH = settings.BASE_DIR / "config" / "personal_companies.json"
PERSONAL_EMAIL = "yashuyashwanth9346@gmail.com"


def run_personal_pipeline() -> bool:
    logger.info("Starting Personal Job Aggregator Pipeline (Hyderabad)...")

    if not PERSONAL_COMPANIES_PATH.exists():
        logger.error(f"Personal companies JSON not found at {PERSONAL_COMPANIES_PATH}")
        return False

    with open(PERSONAL_COMPANIES_PATH, "r") as f:
        companies = json.load(f)

    logger.info(f"Loaded {len(companies)} target companies for personal pipeline.")

    raw_jobs: List[Dict[str, Any]] = []

    for comp in companies:
        name = comp.get("name")
        ats_type = comp.get("ats", "").lower()
        token = comp.get("token")
        careers_url = comp.get("careers_url", "")

        scraper = None
        if ats_type == "greenhouse":
            scraper = GreenhouseScraper(name, token, careers_url)
        elif ats_type == "ashby":
            scraper = AshbyScraper(name, token, careers_url)
        else:
            continue

        try:
            company_jobs = scraper.scrape()
            raw_jobs.extend(company_jobs)
        except Exception as e:
            logger.error(f"Failed to scrape {name}: {str(e)}", exc_info=True)

    logger.info(f"Collected {len(raw_jobs)} raw jobs from all scrapers.")

    passed: List[Dict[str, Any]] = []
    seen_links = set()

    for job in raw_jobs:
        link = job.get("apply_link", "").lower()
        if link in seen_links:
            continue

        is_match, reason, enriched = filter_personal_job(job)
        if is_match:
            passed.append(enriched)
            seen_links.add(link)

    logger.info(f"After filter: {len(passed)} matching jobs (Hyderabad + role + experience).")

    if not passed:
        logger.warning("No matching jobs found today for personal pipeline.")
        return True

    sorted_jobs = sorted(passed, key=lambda j: j.get("date_posted", "") or "0000-00-00", reverse=True)
    final_selection = sorted_jobs[:50]
    logger.info(f"Selected top {len(final_selection)} jobs for personal report.")

    try:
        excel_path = generate_personal_excel(final_selection)
    except Exception as e:
        logger.error(f"Error creating personal Excel report: {str(e)}", exc_info=True)
        return False

    try:
        send_personal_email(excel_path, final_selection)
    except Exception as e:
        logger.error(f"Error sending personal email: {str(e)}", exc_info=True)

    logger.info("Personal pipeline executed successfully.")
    return True


def generate_personal_excel(jobs: List[Dict[str, Any]]) -> str:
    import datetime
    from pathlib import Path
    import pandas as pd
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    today_str = datetime.date.today().strftime("%d%m%Y")
    base_filename = f"PersonalJobs_{today_str}"
    personal_dir = settings.DATA_DIR / "personal"
    personal_dir.mkdir(parents=True, exist_ok=True)
    file_path = personal_dir / f"{base_filename}.xlsx"

    counter = 2
    while file_path.exists():
        try:
            with file_path.open("r+b"):
                pass
            break
        except PermissionError:
            file_path = personal_dir / f"{base_filename}_v{counter}.xlsx"
            counter += 1

    data = []
    for idx, job in enumerate(jobs, 1):
        data.append({
            "S.No": idx,
            "Date Added": job.get("date_posted") or datetime.date.today().strftime("%Y-%m-%d"),
            "Company": job.get("company", ""),
            "Job Title": job.get("title", ""),
            "Location": job.get("location", ""),
            "Experience": job.get("experience_metadata", "Not Specified"),
            "Apply Link": job.get("apply_link", ""),
        })

    df = pd.DataFrame(data)
    writer = pd.ExcelWriter(file_path, engine="openpyxl")
    df.to_excel(writer, index=False, startrow=3, sheet_name="Jobs")

    ws = writer.sheets["Jobs"]

    font_family = "Segoe UI"
    brand_color = "2E4057"
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color=brand_color, end_color=brand_color, fill_type="solid")
    cell_font = Font(name=font_family, size=10)
    link_font = Font(name=font_family, size=10, color="0563C1", underline="single")
    zebra_fill = PatternFill(start_color="F2F6FA", end_color="F2F6FA", fill_type="solid")
    white_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    thin_side = Side(border_style="thin", color="D3D3D3")
    cell_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")

    ws["A1"] = "Full Stack / AI-ML Jobs - Hyderabad (Fresher to 3 Yrs)"
    ws["A1"].font = Font(name=font_family, size=16, bold=True, color=brand_color)

    generated_time = datetime.datetime.now().strftime("%B %d, %Y - %I:%M %p")
    ws["A2"] = f"Report: {generated_time} | Total: {len(jobs)}"
    ws["A2"].font = Font(name=font_family, size=10, italic=True, color="595959")

    ws.row_dimensions[1].height = 25
    ws.row_dimensions[2].height = 18
    ws.row_dimensions[4].height = 28

    num_cols = len(df.columns)
    for col_idx in range(1, num_cols + 1):
        cell = ws.cell(row=4, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align if col_idx <= 2 else left_align
        cell.border = cell_border

    for row_idx in range(5, 5 + len(df)):
        ws.row_dimensions[row_idx].height = 20
        is_even = row_idx % 2 == 0
        fill = zebra_fill if is_even else white_fill
        for col_idx in range(1, num_cols + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.font = cell_font
            cell.fill = fill
            cell.border = cell_border
            cell.alignment = center_align if col_idx <= 2 else left_align
            if col_idx == 7:
                link_url = cell.value
                if link_url and str(link_url).startswith("http"):
                    cell.hyperlink = link_url
                    cell.value = link_url
                    cell.font = link_font

    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_len = max((len(str(cell.value or "")) for cell in col if cell.row >= 4), default=10)
        ws.column_dimensions[col_letter].width = max(max_len + 4, 10)

    writer.close()
    logger.info(f"Personal Excel generated: {file_path}")
    return str(file_path.resolve())


def send_personal_email(excel_path: str, jobs: List[Dict[str, Any]]) -> bool:
    import datetime
    import smtplib
    from email.mime.multipart import MIMEMultipart
    from email.mime.text import MIMEText
    from email.mime.base import MIMEBase
    from email import encoders
    from pathlib import Path

    if not settings.SMTP_USER or not settings.SMTP_PASSWORD:
        logger.warning("SMTP credentials missing. Personal email skipped.")
        return False

    excel_file = Path(excel_path)
    if not excel_file.exists():
        return False

    today_str = datetime.date.today().strftime("%d/%m/%Y")
    subject = f"Daily Job Leads ({today_str}) - {len(jobs)} Hyderabad Jobs"

    msg = MIMEMultipart()
    msg["From"] = settings.EMAIL_FROM
    msg["To"] = PERSONAL_EMAIL
    msg["Subject"] = subject

    html = f"""
    <div style="font-family: Segoe UI, sans-serif; max-width: 600px; margin: auto;">
        <h2 style="color: #2E4057;">Daily Job Leads - Hyderabad</h2>
        <p>Hi Yashwanth, today <strong>{len(jobs)}</strong> matching jobs found.</p>
        <p>Excel sheet attached with full details + apply links.</p>
        <hr>
        <p style="color: #666; font-size: 12px;">Full Stack / AI-ML | Fresher to 3 Yrs | Hyderabad/India</p>
    </div>
    """
    msg.attach(MIMEText(html, "html"))

    with open(excel_file, "rb") as attachment:
        part = MIMEBase("application", "octet-stream")
        part.set_payload(attachment.read())
        encoders.encode_base64(part)
        part.add_header("Content-Disposition", f"attachment; filename= {excel_file.name}")
        msg.attach(part)

    server = smtplib.SMTP(settings.SMTP_HOST, settings.SMTP_PORT)
    server.starttls()
    server.login(settings.SMTP_USER, settings.SMTP_PASSWORD)
    server.sendmail(settings.EMAIL_FROM, [PERSONAL_EMAIL], msg.as_string())
    server.quit()

    logger.info(f"Personal email sent to {PERSONAL_EMAIL}")
    return True
